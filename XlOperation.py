# -*- coding: utf-8 -*-
"""
Created on Wed Jan 30 14:23:35 2019

@author: SKourav
"""
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws['A1'] = 42

ws.append([1,2,3,4])

import datetime

ws['A2'] = datetime.datetime.now()

wb.save("pythonExcel.xlsx")